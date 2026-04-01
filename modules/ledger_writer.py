"""
ledger_writer.py — SR 11-7 Validation Ledger Writer

Architecture: Heterogeneous-tab strategy.
- Each module's DataFrames are written to individual named tabs preserving
  all domain-specific columns (baseline_value, shocked_value, variance_pct,
  extracted_value, truth_value, missing_count, etc.).
- The unifying contract: every tab contains a `state` column.
- openpyxl dynamically hunts the `state` column in each sheet and applies
  conditional formatting (Green / Red / Amber) regardless of its position.
- NO canonical schema flattening. No information is lost.
- Tab 00_Executive_Summary is generated dynamically and always written first.

I/O pattern: single-pass in-memory.
  pd.ExcelWriter(engine="openpyxl") holds the openpyxl workbook in memory.
  writer.sheets[sheet_name] exposes the live worksheet object immediately
  after to_excel(), so all formatting is applied in-memory before the
  context manager flushes exactly one write to disk on exit.
"""

from __future__ import annotations

import re
import warnings
from datetime import datetime, timezone
from typing import Dict, List, Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour constants  (openpyxl ARGB hex — no leading #)
# ---------------------------------------------------------------------------
FILL_PASS   = PatternFill("solid", fgColor="C6EFCE")   # Excel "Good"    green
FILL_FAIL   = PatternFill("solid", fgColor="FFC7CE")   # Excel "Bad"     red
FILL_NA     = PatternFill("solid", fgColor="FFEB9C")   # Excel "Neutral" amber
FILL_HEADER = PatternFill("solid", fgColor="1F3864")   # Deep navy

FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
FONT_BODY   = Font(name="Calibri", size=10)

_THIN = Side(style="thin", color="D9D9D9")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ILLEGAL_CHARS = re.compile(r'[\\/*?:\[\]]')

# Expanded vocabulary: validation (PASS/FAIL/N/A) + elasticity (STABLE/SENSITIVE/BREACH)
_STATE_FILLS: dict[str, PatternFill] = {
    "PASS":      FILL_PASS,
    "STABLE":    FILL_PASS,
    "FAIL":      FILL_FAIL,
    "BREACH":    FILL_FAIL,
    "N/A":       FILL_NA,
    "SENSITIVE": FILL_NA,
}

# The key injected by write_ledger — excluded from pre-flight state check
_SUMMARY_KEY = "00_Executive_Summary"

# Risk columns tried in priority order when computing stress severity
_RISK_COL_PRIORITY = ("max_abs_delta_pct", "delta_pct", "variance_pct")


# ---------------------------------------------------------------------------
# Summary generator
# ---------------------------------------------------------------------------

def _generate_executive_summary(
    flat_results: Dict[str, pd.DataFrame],
) -> pd.DataFrame:
    """
    Build the executive summary DataFrame for the 00_Executive_Summary tab.

    Computes three SR 11-7 headline metrics from the flat results dict:

    1. Pipeline Health (Data_Integrity__ prefix)
       total checks run vs. number of FAIL states.

    2. AI Hallucination Rate (AI_Audit__ prefix)
       FAIL % of audited numeric claims (denominator = PASS + FAIL only;
       N/A rows are qualitative claims and excluded from the rate).

    3. Top Economic Risks (Stress_Testing__ prefix)
       Top-3 rows by absolute risk_pct, derived from the first available
       column in: max_abs_delta_pct → delta_pct → variance_pct.
       Includes assumption name and state label.

    Parameters
    ----------
    flat_results : dict[str, pd.DataFrame]
        Flat results dict from flatten_results().  The summary key
        ``00_Executive_Summary`` must NOT be present (pre-injection state).

    Returns
    -------
    pd.DataFrame
        Columns: ["Metric", "Value", "Notes"]
        Does not include a ``state`` column — the summary tab is exempt from
        the state contract enforced on module tabs.
    """
    rows: list[dict] = []

    # --- Metric 1: Pipeline Health ------------------------------------------
    integrity_dfs = [
        df for key, df in flat_results.items()
        if key.startswith("Data_Integrity__")
    ]
    if integrity_dfs:
        combined = pd.concat(integrity_dfs, ignore_index=True)
        total_checks = len(combined)
        fail_checks  = int((combined["state"] == "FAIL").sum())
        pass_checks  = total_checks - fail_checks
        health_label = "PASS" if fail_checks == 0 else "FAIL"
        rows.append({
            "Metric": "Pipeline Health",
            "Value":  health_label,
            "Notes":  (
                f"{pass_checks}/{total_checks} checks passed; "
                f"{fail_checks} FAIL(s) detected across data_integrity module."
            ),
        })
    else:
        rows.append({
            "Metric": "Pipeline Health",
            "Value":  "N/A",
            "Notes":  "No Data_Integrity__ results found.",
        })

    # --- Metric 2: AI Hallucination Rate ------------------------------------
    audit_dfs = [
        df for key, df in flat_results.items()
        if key.startswith("AI_Audit__")
    ]
    if audit_dfs:
        combined = pd.concat(audit_dfs, ignore_index=True)
        # Only audited numeric claims count — exclude qualitative N/A rows
        numeric_mask = combined["state"].isin(["PASS", "FAIL"])
        numeric_rows = combined[numeric_mask]
        audited      = len(numeric_rows)
        failed       = int((numeric_rows["state"] == "FAIL").sum())
        if audited > 0:
            rate_decimal = round(failed / audited, 3)  # Store as decimal (e.g., 0.667)
            notes_str = (
                f"{failed}/{audited} audited numeric claims failed tolerance check. "
                f"N/A rows (qualitative) excluded from denominator."
            )
        else:
            rate_decimal = 0.0
            notes_str = "No auditable numeric claims found."
        rows.append({
            "Metric": "AI Hallucination Rate",
            "Value":  rate_decimal,  # Store as float, not string
            "Notes":  notes_str,
        })
    else:
        rows.append({
            "Metric": "AI Hallucination Rate",
            "Value":  "N/A",
            "Notes":  "No AI_Audit__ results found.",
        })

    # --- Metric 3: Top Economic Risks ---------------------------------------
    stress_dfs = [
        df for key, df in flat_results.items()
        if key.startswith("Stress_Testing__")
    ]
    if stress_dfs:
        risk_records: list[dict] = []
        for df in stress_dfs:
            # Find the first available risk column
            risk_col = next(
                (c for c in _RISK_COL_PRIORITY if c in df.columns), None
            )
            if risk_col is None:
                continue
            for _, row in df.iterrows():
                val = row[risk_col]
                try:
                    abs_val = abs(float(val))
                except (TypeError, ValueError):
                    continue
                assumption = row.get("assumption", "unknown")
                state      = row.get("state", "?")
                risk_records.append({
                    "assumption": assumption,
                    "abs_risk":   abs_val,
                    "risk_pct":   float(val),
                    "state":      state,
                    "risk_col":   risk_col,
                })

        if risk_records:
            risk_df = pd.DataFrame(risk_records)
            # Sort descending, then keep only the single worst shock per
            # assumption so one assumption cannot dominate all three slots.
            top3 = (
                risk_df
                .sort_values("abs_risk", ascending=False)
                .reset_index(drop=True)
                .drop_duplicates(subset=["assumption"], keep="first")
                .head(3)
            )
            top_strs = [
                f"{r['assumption']} ({r['risk_pct']:+.2f}% [{r['state']}])"
                for _, r in top3.iterrows()
            ]
            rows.append({
                "Metric": "Top Economic Risks",
                "Value":  top3.iloc[0]["state"],   # worst single-assumption state
                "Notes":  "; ".join(top_strs),
            })
        else:
            rows.append({
                "Metric": "Top Economic Risks",
                "Value":  "N/A",
                "Notes":  "No risk columns found in Stress_Testing__ results.",
            })
    else:
        rows.append({
            "Metric": "Top Economic Risks",
            "Value":  "N/A",
            "Notes":  "No Stress_Testing__ results found.",
        })

    # --- Run timestamp ------------------------------------------------------
    rows.append({
        "Metric": "Run Timestamp (UTC)",
        "Value":  datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "Notes":  "Timestamp of this validation run.",
    })

    return pd.DataFrame(rows, columns=["Metric", "Value", "Notes"])


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

def _sanitize_sheet_name(name: str) -> str:
    """Excel tab names: max 31 chars, no illegal characters."""
    return _ILLEGAL_CHARS.sub("_", name)[:31]


def _find_state_col_index(ws) -> Optional[int]:
    """
    Return the 1-based column index of the first header cell whose value is
    exactly 'state'.  Returns None if not found.

    Uses exact equality (== "state") to avoid accidentally matching
    'state_type', which also contains the word 'state'.
    """
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).strip() == "state":
            return cell.column
    return None


def _apply_header_style(ws) -> None:
    """Navy bold header row + auto-fit column widths."""
    for cell in ws[1]:
        cell.font      = FONT_HEADER
        cell.fill      = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=8,
        )
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max_len + 4, 60)


def _apply_body_style(ws) -> None:
    """Thin borders + Calibri 10 on every data row."""
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font      = FONT_BODY
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def _apply_state_formatting(ws, state_col_idx: int) -> None:
    """
    Fill only the state cell in each data row based on its value.

    Validation vocabulary:
        PASS      -> green  (#C6EFCE)
        FAIL      -> red    (#FFC7CE)
        N/A       -> amber  (#FFEB9C)

    Elasticity vocabulary:
        STABLE    -> green  (#C6EFCE)
        SENSITIVE -> amber  (#FFEB9C)
        BREACH    -> red    (#FFC7CE)

    Every other column is untouched.  Heterogeneous numeric data is
    preserved exactly as written.
    """
    for row in ws.iter_rows(
        min_row=2, min_col=state_col_idx, max_col=state_col_idx
    ):
        for cell in row:
            val  = str(cell.value).strip().upper() if cell.value is not None else ""
            fill = _STATE_FILLS.get(val)
            if fill:
                cell.fill = fill


def _freeze_header(ws) -> None:
    """Freeze row 1 so the header stays visible while scrolling."""
    ws.freeze_panes = "A2"


def _apply_summary_formatting(ws) -> None:
    """
    Apply percentage number format to the AI Hallucination Rate value cell.
    
    The summary sheet stores AI Hallucination Rate as a decimal float
    (e.g., 0.667) which needs to be displayed as a percentage (66.7%).
    This prevents Excel's yellow triangle warning about text-formatted numbers.
    """
    # Find the AI Hallucination Rate row and apply percentage format
    for row in ws.iter_rows(min_row=2):  # Skip header row
        metric_cell = row[0]  # Column A (Metric)
        value_cell = row[1]   # Column B (Value)
        
        if metric_cell.value == "AI Hallucination Rate":
            # Apply percentage format: 0.667 → 66.7%
            if isinstance(value_cell.value, (int, float)):
                value_cell.number_format = '0.0%'
            break


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_master_results(
    data_integrity_results: Dict[str, pd.DataFrame],
    stress_test_results: Dict[str, pd.DataFrame],
    ai_audit_results: Dict[str, pd.DataFrame],
) -> Dict[str, Dict[str, pd.DataFrame]]:
    """
    Bundle all module outputs into a single nested result object.

    Parameters
    ----------
    data_integrity_results : dict[str, pd.DataFrame]
        Output of run_integrity_checks().
    stress_test_results : dict[str, pd.DataFrame]
        Output of run_stress_tests().
    ai_audit_results : dict[str, pd.DataFrame]
        Output of evaluate_claims().

    Returns
    -------
    dict[str, dict[str, pd.DataFrame]]
        Nested structure keyed by module name.
    """
    return {
        "Data_Integrity": data_integrity_results,
        "Stress_Testing": stress_test_results,
        "AI_Audit": ai_audit_results,
    }


def flatten_results(
    master_results: Dict[str, Dict[str, pd.DataFrame]],
) -> Dict[str, pd.DataFrame]:
    """
    Flatten nested module results into sheet-friendly names.

    Parameters
    ----------
    master_results : dict[str, dict[str, pd.DataFrame]]
        Output of build_master_results().

    Returns
    -------
    dict[str, pd.DataFrame]
        Keys are ``ModuleName__check_name`` (e.g. ``Data_Integrity__missing``).
    """
    flat: Dict[str, pd.DataFrame] = {}
    for module_name, module_results in master_results.items():
        for check_name, df in module_results.items():
            flat[f"{module_name}__{check_name}"] = df
    return flat


def write_ledger(
    flat_results: Dict[str, pd.DataFrame],
    output_path: str = "sr117_validation_ledger.xlsx",
    *,
    sheet_order: Optional[List[str]] = None,
) -> str:
    """
    Write the SR 11-7 Validation Ledger to a formatted Excel workbook.

    Always writes ``00_Executive_Summary`` as the first tab, generated
    dynamically from ``flat_results`` at call time.

    Parameters
    ----------
    flat_results : dict[str, pd.DataFrame]
        Output of flatten_results().  Keys become sheet names.

        Contract: every DataFrame MUST contain a ``state`` column.
        Validation vocabulary: PASS / FAIL / N/A
        Elasticity vocabulary: STABLE / SENSITIVE / BREACH
        All other columns are written verbatim.

    output_path : str
        Destination .xlsx path.

    sheet_order : list[str] | None
        Optional explicit tab ordering for module sheets.  The summary tab
        is always prepended regardless.  Keys absent from sheet_order are
        appended alphabetically after any explicitly ordered keys.

    Returns
    -------
    str
        Resolved output path (identical to ``output_path``).

    Raises
    ------
    ValueError
        If any DataFrame is missing the mandatory ``state`` column.
    """
    # Pre-flight: verify state contract across module sheets -----------------
    # (00_Executive_Summary is injected below and is exempt from this check)
    for sheet_key, df in flat_results.items():
        if sheet_key == _SUMMARY_KEY:
            continue
        if "state" not in df.columns:
            raise ValueError(
                f"Sheet '{sheet_key}' is missing the mandatory `state` column. "
                "All module DataFrames must carry a state column "
                "(PASS / FAIL / N/A or STABLE / SENSITIVE / BREACH) "
                "before being passed to write_ledger()."
            )

    # Shallow copy to avoid mutating the caller's dict -----------------------
    results_to_write = dict(flat_results)

    # Inject the executive summary -------------------------------------------
    results_to_write[_SUMMARY_KEY] = _generate_executive_summary(flat_results)

    # Determine write order: summary tab is ALWAYS first ---------------------
    if sheet_order is None:
        module_keys = sorted(
            k for k in results_to_write if k != _SUMMARY_KEY
        )
    else:
        seen: set = set()
        module_keys = []
        for k in sheet_order:
            if k in results_to_write and k != _SUMMARY_KEY:
                module_keys.append(k)
                seen.add(k)
        for k in sorted(results_to_write.keys()):
            if k not in seen and k != _SUMMARY_KEY:
                module_keys.append(k)

    ordered_keys = [_SUMMARY_KEY] + module_keys

    # Single-pass write and format -------------------------------------------
    # pd.ExcelWriter(engine="openpyxl") holds the workbook in memory.
    # writer.sheets[sheet_name] exposes the live openpyxl worksheet object
    # immediately after to_excel(), so all formatting runs in-memory.
    # The context manager executes exactly one disk write on exit.
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for key in ordered_keys:
            sheet_name = _sanitize_sheet_name(key)

            # 1. Dump DataFrame to the in-memory workbook
            results_to_write[key].to_excel(writer, sheet_name=sheet_name, index=False)

            # 2. Grab the live openpyxl worksheet object directly from the writer
            ws = writer.sheets[sheet_name]

            # 3. Apply all openpyxl formatting in memory
            _apply_header_style(ws)
            _apply_body_style(ws)
            _freeze_header(ws)
            
            # 3a. Summary-specific formatting (percentage number format)
            if key == _SUMMARY_KEY:
                _apply_summary_formatting(ws)

            # 4. State-column colouring — skipped gracefully for summary tab
            state_col_idx = _find_state_col_index(ws)
            if state_col_idx is not None:
                _apply_state_formatting(ws, state_col_idx)
            elif key != _SUMMARY_KEY:
                # Only warn for module tabs that should have a state column
                warnings.warn(
                    f"Sheet '{sheet_name}': `state` column not found — "
                    "conditional formatting skipped for this sheet.",
                    RuntimeWarning,
                    stacklevel=2,
                )

    # Context manager exits here — single disk write. -------------------------
    return output_path
