"""
tests/test_ledger_writer.py

Integration tests for modules/ledger_writer.py.

Strategy: call write_ledger() with synthetic flat_results, then reopen the
produced .xlsx with openpyxl and assert workbook structure, tab ordering,
executive summary metric values, conditional fills, and mutation safety.

No monkeypatching is needed — ledger_writer.py reads only the DataFrames
passed to it; it never calls get_domain_config().

Coverage:
  1. Workbook is created; 00_Executive_Summary is the first tab.
  2. Executive summary metrics are correct (labels, values, notes).
  3. State formatting is applied on raw tabs; summary tab is skipped safely.
  4. write_ledger does not mutate the caller's flat_results dict.
  5. A Stress_Testing__ sheet with no recognized risk column is skipped cleanly.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from modules.ledger_writer import write_ledger


# ---------------------------------------------------------------------------
# Synthetic fixtures — module-level constants, fully deterministic
# ---------------------------------------------------------------------------

_DI_MISSING = pd.DataFrame([
    {
        "column": "revenue", "missing_count": 1, "missing_pct": 20.0,
        "state": "FAIL", "state_type": "validation",
    },
    {
        "column": "year", "missing_count": 0, "missing_pct": 0.0,
        "state": "PASS", "state_type": "validation",
    },
])
# Pipeline Health → 1 FAIL / 2 total → value "FAIL"

_AI_AUDIT = pd.DataFrame([
    {
        "claim_text": "Revenue was $45M.", "source_key": "revenue",
        "unit": "usd", "confidence": 1.0,
        "extracted_value": 45.0, "source_value": 45.0,
        "variance": 0.0, "variance_pct": 0.0,
        "notes": "Passed.", "state": "PASS", "state_type": "validation",
    },
    {
        "claim_text": "Revenue was $50M.", "source_key": "revenue",
        "unit": "usd", "confidence": 0.9,
        "extracted_value": 50.0, "source_value": 45.0,
        "variance": 5.0, "variance_pct": 11.11,
        "notes": "Failed tolerance.", "state": "FAIL", "state_type": "validation",
    },
    {
        "claim_text": "Strong brand moat.", "source_key": None,
        "unit": None, "confidence": None,
        "extracted_value": None, "source_value": None,
        "variance": None, "variance_pct": None,
        "notes": "Qualitative claim.", "state": "N/A", "state_type": "validation",
    },
])
# AI Hallucination Rate → 1 FAIL / 2 auditable (PASS+FAIL, N/A excluded) = 50.0%

_STRESS_SUMMARY = pd.DataFrame([
    {
        "assumption": "free_cash_flow",
        "base_output": 1000.0, "output_minus_1sd": 960.0, "output_plus_1sd": 1040.0,
        "delta_pct_minus": -4.0, "delta_pct_plus": 4.0, "max_abs_delta_pct": 4.0,
        "state": "STABLE", "state_type": "elasticity",
    },
    {
        "assumption": "growth_rate",
        "base_output": 1000.0, "output_minus_1sd": 890.0, "output_plus_1sd": 1110.0,
        "delta_pct_minus": -11.0, "delta_pct_plus": 11.0, "max_abs_delta_pct": 11.0,
        "state": "SENSITIVE", "state_type": "elasticity",
    },
    {
        "assumption": "discount_rate",
        "base_output": 1000.0, "output_minus_1sd": 750.0, "output_plus_1sd": 1300.0,
        "delta_pct_minus": -25.0, "delta_pct_plus": 30.0, "max_abs_delta_pct": 30.0,
        "state": "BREACH", "state_type": "elasticity",
    },
])
# Top risk → discount_rate, abs = 30.0, state = "BREACH"

# Stress sheet with NO recognised risk column (max_abs_delta_pct / delta_pct / variance_pct)
_STRESS_AUX = pd.DataFrame([
    {"assumption": "aux_factor", "state": "STABLE", "state_type": "elasticity"},
])

_FLAT_RESULTS: dict[str, pd.DataFrame] = {
    "Data_Integrity__missing":        _DI_MISSING,
    "AI_Audit__claim_audit":          _AI_AUDIT,
    "Stress_Testing__one_sd_summary": _STRESS_SUMMARY,
}


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def _get_summary_metrics(ws) -> dict[str, dict[str, object]]:
    """
    Return {Metric: {"value": ..., "notes": ...}} from the summary sheet.
    Row 1 is the header; rows 2+ are data.  None-keyed rows are skipped.
    """
    return {
        row[0]: {"value": row[1], "notes": row[2]}
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    }


def _find_state_col(ws) -> int | None:
    """Return 1-based column index of the exact 'state' header, or None."""
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).strip() == "state":
            return cell.column
    return None


# ---------------------------------------------------------------------------
# Test 1 — Workbook created; 00_Executive_Summary is first tab
# ---------------------------------------------------------------------------

def test_write_ledger_creates_workbook_and_summary_first(tmp_path: Path):
    """
    write_ledger must produce a .xlsx file whose first sheet is exactly
    '00_Executive_Summary', with all module tabs also present.
    """
    out = tmp_path / "ledger.xlsx"
    result = write_ledger(dict(_FLAT_RESULTS), output_path=str(out))

    assert out.exists(), "write_ledger must create the output file"
    assert result == str(out), "write_ledger must return the output path"

    wb = openpyxl.load_workbook(str(out))
    assert wb.sheetnames[0] == "00_Executive_Summary", (
        f"First tab must be '00_Executive_Summary'; got {wb.sheetnames[0]!r}"
    )
    assert "Data_Integrity__missing"        in wb.sheetnames
    assert "AI_Audit__claim_audit"          in wb.sheetnames
    assert "Stress_Testing__one_sd_summary" in wb.sheetnames


# ---------------------------------------------------------------------------
# Test 2 — Executive summary metric labels and values
# ---------------------------------------------------------------------------

def test_executive_summary_contains_expected_metrics(tmp_path: Path):
    """
    The 00_Executive_Summary sheet must have:
      - header row ["Metric", "Value", "Notes"]
      - rows for Pipeline Health, AI Hallucination Rate, Top Economic Risks,
        Run Timestamp (UTC)
      - AI hallucination rate value containing "50.0" (1/2 auditable)
      - Top Economic Risks:
          Value  = "BREACH"   (state of worst assumption)
          Notes  contains "discount_rate"
    """
    out = tmp_path / "ledger.xlsx"
    write_ledger(dict(_FLAT_RESULTS), output_path=str(out))

    wb = openpyxl.load_workbook(str(out))
    ws = wb["00_Executive_Summary"]

    # Header row
    headers = [ws.cell(1, c).value for c in range(1, 4)]
    assert headers == ["Metric", "Value", "Notes"], (
        f"Summary header must be ['Metric', 'Value', 'Notes']; got {headers!r}"
    )

    metrics = _get_summary_metrics(ws)

    # All expected metric labels present
    for label in ("Pipeline Health", "AI Hallucination Rate",
                  "Top Economic Risks", "Run Timestamp (UTC)"):
        assert label in metrics, f"Expected metric row '{label}' missing from summary"

    # AI Hallucination Rate: 1 FAIL / 2 auditable = 50.0% = 0.5 decimal
    # Value is now stored as a float (0.5), not a string ("50.0%")
    hallucination_value = metrics["AI Hallucination Rate"]["value"]
    assert isinstance(hallucination_value, (int, float)), (
        f"Expected hallucination rate to be numeric; got {type(hallucination_value)}"
    )
    assert abs(hallucination_value - 0.5) < 0.01, (
        f"Expected hallucination rate ~0.5 (50%); got {hallucination_value!r}"
    )

    # Top Economic Risks: single row (not three separate rows)
    # Value = worst assumption's state; Notes = semicolon-joined top-3 assumptions
    assert metrics["Top Economic Risks"]["value"] == "BREACH", (
        f"Expected Top Economic Risks Value='BREACH'; got {metrics['Top Economic Risks']['value']!r}"
    )
    assert "discount_rate" in str(metrics["Top Economic Risks"]["notes"]), (
        f"'discount_rate' must appear in Top Economic Risks Notes; "
        f"got {metrics['Top Economic Risks']['notes']!r}"
    )


# ---------------------------------------------------------------------------
# Test 3 — State formatting applied on raw tabs; summary tab skipped safely
# ---------------------------------------------------------------------------

def test_state_formatting_skips_summary_and_formats_raw_tabs(tmp_path: Path):
    """
    - Summary tab has no 'state' column and raises no error.
    - A raw tab with a 'state' column has non-default fills on state cells.
    - Freeze panes (A2) are set on both summary and raw tabs.
    """
    out = tmp_path / "ledger.xlsx"
    write_ledger(dict(_FLAT_RESULTS), output_path=str(out))

    wb = openpyxl.load_workbook(str(out))

    # --- Summary tab: no 'state' column ---
    ws_summary = wb["00_Executive_Summary"]
    summary_headers = [
        ws_summary.cell(1, c).value
        for c in range(1, ws_summary.max_column + 1)
    ]
    assert "state" not in summary_headers, (
        "Summary sheet must not have a 'state' column"
    )

    # --- Raw tab: AI_Audit__claim_audit ---
    ws_raw = wb["AI_Audit__claim_audit"]
    state_col = _find_state_col(ws_raw)
    assert state_col is not None, "Raw sheet must have a 'state' header column"

    # Collect (state_value, fill_rgb) for all data rows
    fills_by_state: dict[str, str] = {}
    for row in ws_raw.iter_rows(min_row=2):
        cell = row[state_col - 1]  # state_col is 1-based
        val = str(cell.value).strip() if cell.value is not None else ""
        if val:
            fills_by_state[val] = cell.fill.fgColor.rgb

    # PASS, FAIL, N/A must all have non-default fills applied
    default_rgb = "00000000"
    for state_val in ("PASS", "FAIL", "N/A"):
        assert state_val in fills_by_state, (
            f"No data row with state='{state_val}' found in AI_Audit__claim_audit"
        )
        assert fills_by_state[state_val] != default_rgb, (
            f"state='{state_val}' cell has default (no-fill) fgColor rgb; "
            "conditional formatting was not applied"
        )

    # Exact fill values (stronger: proves correct colour per vocabulary)
    assert fills_by_state["PASS"] == "00C6EFCE", (
        f"PASS fill expected '00C6EFCE'; got {fills_by_state['PASS']!r}"
    )
    assert fills_by_state["FAIL"] == "00FFC7CE", (
        f"FAIL fill expected '00FFC7CE'; got {fills_by_state['FAIL']!r}"
    )
    assert fills_by_state["N/A"] == "00FFEB9C", (
        f"N/A fill expected '00FFEB9C'; got {fills_by_state['N/A']!r}"
    )

    # Stress tab: STABLE/SENSITIVE/BREACH fills
    ws_stress = wb["Stress_Testing__one_sd_summary"]  # 30 chars — no truncation needed
    stress_state_col = _find_state_col(ws_stress)
    assert stress_state_col is not None

    stress_fills: dict[str, str] = {}
    for row in ws_stress.iter_rows(min_row=2):
        cell = row[stress_state_col - 1]
        val = str(cell.value).strip() if cell.value is not None else ""
        if val:
            stress_fills[val] = cell.fill.fgColor.rgb

    assert stress_fills.get("STABLE")    == "00C6EFCE"
    assert stress_fills.get("SENSITIVE") == "00FFEB9C"
    assert stress_fills.get("BREACH")    == "00FFC7CE"

    # Freeze panes on both tabs
    assert ws_summary.freeze_panes == "A2", "Summary freeze_panes must be 'A2'"
    assert ws_raw.freeze_panes     == "A2", "Raw tab freeze_panes must be 'A2'"


# ---------------------------------------------------------------------------
# Test 4 — Mutation guard: write_ledger must not modify caller's dict
# ---------------------------------------------------------------------------

def test_write_ledger_does_not_mutate_input_flat_results(tmp_path: Path):
    """
    write_ledger uses dict(flat_results) internally.  The caller's original
    dict must not gain the '00_Executive_Summary' key or lose any keys.
    """
    caller_dict = dict(_FLAT_RESULTS)   # independent copy
    original_keys = set(caller_dict.keys())

    write_ledger(caller_dict, output_path=str(tmp_path / "ledger.xlsx"))

    assert set(caller_dict.keys()) == original_keys, (
        f"write_ledger mutated the caller's dict; "
        f"new keys: {set(caller_dict.keys()) - original_keys}"
    )
    assert "00_Executive_Summary" not in caller_dict, (
        "write_ledger must not inject '00_Executive_Summary' into the caller's dict"
    )


# ---------------------------------------------------------------------------
# Test 5 — Stress sheet with no recognised risk column is skipped cleanly
# ---------------------------------------------------------------------------

def test_stress_sheet_with_no_recognized_risk_column_is_skipped_cleanly(
    tmp_path: Path,
):
    """
    Including a Stress_Testing__* sheet that has none of
    max_abs_delta_pct / delta_pct / variance_pct must not raise.
    write_ledger succeeds and the executive summary still reflects risk
    data from the sheet that DOES have a risk column.
    """
    results_with_aux = {**_FLAT_RESULTS, "Stress_Testing__aux": _STRESS_AUX}

    out = tmp_path / "ledger.xlsx"
    write_ledger(results_with_aux, output_path=str(out))  # must not raise

    wb = openpyxl.load_workbook(str(out))
    assert "00_Executive_Summary" in wb.sheetnames

    ws = wb["00_Executive_Summary"]
    metrics = _get_summary_metrics(ws)

    # Top Economic Risks must still be populated from one_sd_summary (not N/A)
    top_risks = metrics.get("Top Economic Risks", {})
    assert top_risks.get("value") != "N/A", (
        "Top Economic Risks should not be N/A when one_sd_summary is present"
    )
    assert "discount_rate" in str(top_risks.get("notes", "")), (
        "Top risk from one_sd_summary (discount_rate) must still appear despite aux sheet"
    )
